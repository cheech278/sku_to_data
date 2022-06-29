import requests
import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

def CreateWB(dest_filename):
    wb = Workbook()
    catnames =[]
    cats=LinksCollector('links.txt')
    for b in range (len(cats)):
        catnames.append(re.sub('https://www\.vidaxl\.com/g/\d+/','',cats[b]))
        wb.create_sheet(title=catnames[b])
        wb.save(filename=dest_filename)


def LinksCollector(filename):
    cats = []
    file=open(filename)
    for line in (file):
        cats.append(line)
        cats[cats.index(line)]=cats[cats.index(line)].replace("\n","")
    return cats

def FinalLinkCollector(catname):
    catnames =[]
    cats=LinksCollector('links.txt')
    for b in range (len(cats)):
        catnames.append(re.sub('https://www\.vidaxl\.com/g/\d+/','',cats[b]))
    for l in range(len(cats)):
        link = cats[catnames.index(catname)]
        results=(re.findall('\(.+ Results\)',requests.get(link).text))
        results[0]=re.sub(',',"",results[0])
        results =re.findall('\d+',results[0])
        finallink = link + '?prefn1=localizedOnline&prefn2=localizedSearchable&prefv1=true&prefv2=true&start=0&sz='+results[0] #Получаем финальный линк с нужным колвом вещей
    return finallink


def PayloadLinks(catname):
    finallink = FinalLinkCollector(catname)
    links=re.findall ('<a class="link" href=".+html',requests.get(str(finallink)).text)
    links2 = []
    fullink = []
    for i in range (len(links)):
        links2.append((re.findall('\/e.+html',links[i])))
        payload = str(links2[i])
        payload = payload[2:-2]
        fullink.append('https://www.vidaxl.com'+payload)
    return fullink

def SkuCollector(catname, dest_filename):
    wa= load_workbook(filename=dest_filename)
    sku= []
    newsku = []
    skulinkprice = []
    fullink = PayloadLinks(catname)
    pricelist = []
    wsx=wa.get_sheet_by_name(name=catname)
    wsx['A1']='Sku'
    wsx['B1']='Link'
    wsx['C1']='Price'
    for i in range (len(fullink)):
        newsite = requests.get(fullink[i])
        pricelist.append(re.findall('"price":"\d+',newsite.text))
        price =re.findall('\d+',str(pricelist[i]))
        price = int(price[0])
        sku.append(re.findall('S\w\w:\d+',newsite.text))
        newsku.append(re.findall('\d+',str(sku[i])))
        skulinkprice.append([str(newsku[i][0]),str(fullink[i]), str(price)+'$'])
        wsx['A'+str(i+2)]=str(skulinkprice[i][0])
        wsx['B'+str(i+2)]=str(skulinkprice[i][1])
        wsx['C'+str(i+2)]=str(skulinkprice[i][2])
        print (skulinkprice[i])
        wa.save(filename=dest_filename)




   